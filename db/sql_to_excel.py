import os
import openpyxl
from openpyxl.styles import PatternFill

from config.init_folders import EXCEL_DB_FILE_PATH
from db.controllers import DbDataController


class Columns:
    """Class for columns names"""
    QUESTION_ID = 1
    QUESTION_TOPIC = 2
    QUESTION_TEXT = 3
    ANSWER_TEXT = 4
    ANSWER_CORRECT_COLOR = 'FF92D050'


class ExcelData:
    TOPIC_NAME_COORD = (1, 1)

    def __init__(self):
        self.sheetname = 'data'
        self.workbook = self._init_file()
        self.worksheet = self.workbook[self.sheetname]
        self._clear_excel()

    def _init_file(self):
        if not os.path.exists(EXCEL_DB_FILE_PATH):
            workbook = openpyxl.Workbook()
            workbook.create_sheet(self.sheetname)
            workbook.save(EXCEL_DB_FILE_PATH)
        return openpyxl.load_workbook(filename=EXCEL_DB_FILE_PATH)

    def _clear_excel(self):
        self.worksheet.delete_rows(1, self.worksheet.max_row)
        self.workbook.save(EXCEL_DB_FILE_PATH)

    def export_to_excel(self):
        """Copy data from SQL to excel"""
        data = DbDataController().read()
        for num, each in enumerate(data):
            self.workbook[self.sheetname].cell(row=num + 1, column=Columns.QUESTION_ID).value = each.id
            self.workbook[self.sheetname].cell(row=num + 1, column=Columns.QUESTION_TOPIC).value = each.topic
            self.workbook[self.sheetname].cell(row=num + 1, column=Columns.QUESTION_TEXT).value = each.question
            for answer_num, answer in enumerate(each.answers):
                self.workbook[self.sheetname].cell(
                    row=num + 1,
                    column=Columns.ANSWER_TEXT + answer_num
                ).value = answer.text

                if not answer.is_correct:
                    continue

                self.workbook[self.sheetname].cell(
                    row=num + 1,
                    column=Columns.ANSWER_TEXT + answer_num
                ).fill = (
                    PatternFill(
                        start_color=Columns.ANSWER_CORRECT_COLOR,
                        end_color=Columns.ANSWER_CORRECT_COLOR,
                        fill_type="solid"
                    ))

        self.workbook.save(EXCEL_DB_FILE_PATH)
        self.workbook.close()


excel_data = ExcelData()
