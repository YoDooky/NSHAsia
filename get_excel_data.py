# модуль для поиска данных в базе данных Excel и добавления их в массив
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import sys

import aux_functions
from config import folders
import log
sys.tracebacklimit = 0

file_name = file_init.database_name  # путь к папке со всеми файлами (драйвер хрома,
# база данных и т.п.)


# функция для сбора инфы с темами для которых нужно теорию прокликать всю
def get_spec_themes():
    workbook = load_workbook(filename=file_name)
    sheet = file_init.database_click_sheet
    themes_list = []
    row_numb = 1
    while workbook[sheet].cell(row=row_numb, column=1).value:
        themes_list.append(workbook[sheet].cell(row=row_numb, column=1).value)
        row_numb += 1
    workbook.close()
    return themes_list


# функция для получения списка багнутых вопросов
def get_bugged_data(start_row, columb_numb, single_element=0):
    workbook = load_workbook(filename=file_name)
    sheet = file_init.database_bugged_sheet
    bugged_question_list = []
    row_numb = start_row
    column_numb = columb_numb
    while workbook[sheet].cell(row=row_numb, column=column_numb).value:
        bugged_question_list.append(workbook[sheet].cell(row=row_numb, column=column_numb).value)
        if single_element:
            break
        row_numb += 1
    return bugged_question_list


# получаем действие из столбца с действие
def get_bug_action(question):
    workbook = load_workbook(filename=file_name)
    sheet = file_init.database_bugged_sheet
    row_numb = 15
    column_numb = 2
    action = None
    while workbook[sheet].cell(row=row_numb, column=column_numb).value:
        if workbook[sheet].cell(row=row_numb, column=column_numb).value == question:
            action = workbook[sheet].cell(row=row_numb, column=column_numb+6).value  # получаем индекс действия которое
            # нужно сделать
            break
        row_numb += 1
    return action


class ExcelData:
    def __init__(self, filename, sheetname):
        self.filename = filename
        self.sheetname = sheetname

    def load(self):
        """Метод для загрузки данных с листа Excel"""
        return load_workbook(self.filename)

    def get_rowdata(self, row, start_column=1, workbook=None):
        """Метод для получения всей строки с данными"""
        if not workbook:
            workbook = load_workbook(filename=self.filename)
        row_data = []
        while workbook[self.sheetname].cell(row=row, column=start_column).value:
            value = workbook[self.sheetname].cell(row=row, column=start_column).value
            row_data.append(str(value))
            start_column += 1
        workbook.close()
        return row_data

    def get_columndata(self, column, start_row=1, workbook=None):
        """Метод для получения всего столбца с данными"""
        if not workbook:
            workbook = load_workbook(filename=self.filename)
        column_data = []
        while workbook[self.sheetname].cell(row=start_row, column=column).value:
            value = workbook[self.sheetname].cell(row=start_row, column=column).value
            column_data.append(value)
            start_row += 1
        workbook.close()
        return column_data

    def get_data(self, start_row=2, start_column=2, single=False, workbook=None):
        """Метод для получения данных с Excel, начиная с указанной строки и столбца, и помещения их в словарь (
        ключ - 1й столбец, значение - все значения в строке до первого пустого столбца).
        Получает данные до первой пустой строки"""
        if not workbook:
            workbook = load_workbook(filename=self.filename)
        if single:  # возвращаем значение ячейки
            return workbook[self.sheetname].cell(row=start_row, column=start_column).value
        database = {}  # словарь содержащий вопросы и ответы из базы данных
        while workbook[self.sheetname].cell(row=start_row, column=start_column).value:
            key = str(workbook[self.sheetname].cell(row=start_row, column=start_column).value)
            value = [[each] for each in self.get_rowdata(start_row, start_column + 1, workbook)]
            database[key] = value
            start_row += 1
        workbook.close()
        return database

    def write_data(self, value, start_row, start_column, horizontal=True, workbook=None):
        """Метод для записи данных в excel и цвет. Если horizontal=True то пишем данные горизонтально вниз"""
        if not workbook:
            workbook = load_workbook(filename=self.filename)
        if isinstance(value, list):
            for num, each in enumerate(value):
                row = num if horizontal else 0
                col = 0 if horizontal else num
                workbook[self.sheetname].cell(row=start_row + row, column=start_column + col).value = each
                self.write_color(start_row + row, start_column + col, workbook)
        else:
            workbook[self.sheetname].cell(row=start_row, column=start_column).value = value
            self.write_color(start_row, start_column, workbook)
        aux_functions.database_permission(workbook, self.filename)
        # workbook.save(filename=self.filename)
        workbook.close()

    def write_color(self, row, column, workbook=None, color='green'):
        """Красим ячейку"""
        if color == 'red':
            color_code = 'E06666'
        else:
            color_code = 'D2FEBE'
        if not workbook:
            workbook = load_workbook(filename=self.filename)
        workbook[self.sheetname].cell(row=row, column=column).fill = PatternFill(
            fgColor=color_code, fill_type='solid')
        aux_functions.database_permission(workbook, self.filename)
        workbook.close()

    def del_data(self, row):
        """Метод для удаления данных из excel"""
        workbook = load_workbook(filename=self.filename)
        workbook[self.sheetname].delete_rows(row)
        aux_functions.database_permission(workbook, self.filename)
        # workbook.save(filename=self.filename)
        workbook.close()

    def get_row(self, workbook=None):
        """Метод для поиска пустой строки"""
        if not workbook:
            workbook = load_workbook(filename=self.filename)
        last_row = workbook[self.sheetname].max_row + 1
        while not workbook[self.sheetname].cell(row=last_row, column=2).value:
            last_row = last_row - 1
        return last_row + 1

    def get_column(self, row=None, workbook=None):
        """Метод для поиска пустого столбца"""
        if not workbook:
            workbook = load_workbook(filename=self.filename)
        if not row:
            return workbook[self.sheetname].max_column + 1
        else:
            column = 2
            while workbook[self.sheetname].cell(row, column).value:
                column += 1
            return column

    def find_cell(self, value, min_row=1, min_col=1, workbook=None):
        """Метод для поиска ячейки по значению в заданном диапазоне excel"""
        if not workbook:
            workbook = load_workbook(filename=self.filename)
        max_row = self.get_row()
        max_col = self.get_column()
        # нубский линейный поиск
        for row in range(min_row, max_row+1):
            for col in range(min_col, max_col+1):
                if workbook[self.sheetname].cell(row=row, column=col).value == value:
                    return [row, col]

    def check_access(self):
        """Метод для проверки доступа файла на запись"""


def write_sdo(correct_data, bug_data=None):
    """Функция для записи вопроса и ответа в SDO"""
    # if not bug_data:
    #     bug_data = []
    # right_question, right_answer = [], []
    # for each in correct_data:
    #     if each not in bug_data:
    #         prog_logging.print_log(f'Помечаю вопрос <{each}> в SDO...')
    #         right_question.append(each)
    #         prog_logging.print_log(f'Помечаю ответ на него <{correct_data[each][0]}> в SDO...\n')
    #         right_answer.append(correct_data[each][0])

    # записываем в SDO все ответы кроме багнутых
    if not bug_data:
        bug_data = []
    prog_logging.print_log('Записываю помеченые вопросы и ответы в SDO...')
    sdo_data = ExcelData(filename=file_init.database_name,
                         sheetname=file_init.database_sheet)
    workbook = sdo_data.load()
    sdo_database = sdo_data.get_data(2, 2, False, workbook)
    for each in correct_data:
        if each not in bug_data:
            if each not in sdo_database:  # если такого вопроса еще нет в SDO
                row = sdo_data.get_row(workbook)
                sdo_data.write_data(each, start_row=row, start_column=2, horizontal=True,
                                    workbook=workbook)  # записываем вопрос в SDO
                sdo_data.write_data(correct_data[each][0], start_row=row, start_column=3, horizontal=True,
                                    workbook=workbook)  # записываем ответ в SDO
            else:
                coord = sdo_data.find_cell(each, 1, 1, workbook)  # находим координаты вопроса
                column = sdo_data.get_column(coord[0], workbook)  # находим пустую колонку
                sdo_data.write_data(correct_data[each][0], start_row=coord[0], start_column=column, horizontal=True,
                                    workbook=workbook)  # записываем ответ в SDO
            workbook = sdo_data.load()

    # row = sdo_data.get_row()
    # sdo_data.write_data(right_question, row, 2)  # записываем вопрос в SDO
    # sdo_data.write_data(right_answer, row, 3)  # записываем ответ в SDO
